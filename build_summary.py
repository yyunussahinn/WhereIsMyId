"""
build_summary.py — Where is My Id
────────────────────────────────────────────────────────────────────────────
Excel dosyasındaki tüm sayfa sheet'lerini birleştirerek:
  • "Data"    → ham verinin tamamı (pivot için)
  • "Summary" → sayfa bazlı özet tablo
  • "Task"    → sadece New Status = "ID Eklenecek (Waiting Dev)" olanlar

Çalıştırma:
  python build_summary.py

EXCEL_FILE; komut satırı argümanı, WIMID_EXCEL_FILE env var veya
sabit yol üzerinden belirlenir (öncelik sırası: arg > env > sabit).

NOT: "Data", "Summary" ve "Task" sheet'leri varsa üzerine yazar.
"""

import os
import sys
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import shared as sh

# ── Excel dosya yolu (arg > env > sabit) ─────────────────────────────────────
_DEFAULT_EXCEL = "/Users/yunus.sahin/PycharmProjects/PythonProject/PIA_Elements/Revize_Elements_Report_Android.xlsx"

if len(sys.argv) > 1:
    EXCEL_FILE = sys.argv[1]
else:
    EXCEL_FILE = os.environ.get("WIMID_EXCEL_FILE", _DEFAULT_EXCEL)

# ── Sabitler ──────────────────────────────────────────────────────────────────
SKIP_SHEETS    = {"Data", "Summary", "Task"}
DATA_START_ROW = 3
DATA_COL_COUNT = 9  # ElementID Page Type Label Value AccID Status NewStatus AISuggestion

NEW_STATUS_COL_LTR    = get_column_letter(8)   # H
AI_SUGGESTION_COL_LTR = get_column_letter(9)   # I

NS_WAITING = sh.NS_WAITING
ALL_STATUSES = sh.ALL_STATUSES


def get_new_status_default(status: str) -> str:
    return "" if status == sh.STATUS_UNIQUE else NS_WAITING


def sheet_ref(name: str) -> str:
    """Sheet adını Excel formülü için güvenli hale getirir."""
    special = set(" !@#$%^&*()-+=[]{}|;:,.<>?")
    return f"'{name}'" if any(ch in special for ch in name) else name


# Renk paleti kısayolları
PALETTE          = sh.STATUS_PALETTE
NEW_STATUS_COLOR = sh.NEW_STATUS_COLOR
AI_COLOR         = sh.AI_SUGGESTION_COLOR

THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def fill(hex_c):                    return PatternFill("solid", fgColor=hex_c)
def font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size)

# ─────────────────────────────────────────────────────────────
# 1. Dosyayı aç
# ─────────────────────────────────────────────────────────────
if not os.path.exists(EXCEL_FILE):
    raise FileNotFoundError(f"Dosya bulunamadı: {EXCEL_FILE}")

wb = openpyxl.load_workbook(EXCEL_FILE)

page_sheets = [s for s in wb.sheetnames if s not in SKIP_SHEETS]
print(f"📂 {len(page_sheets)} sayfa sheet'i bulundu: {', '.join(page_sheets)}")

# ─────────────────────────────────────────────────────────────
# 2. Tüm sheet'lerden veriyi oku
# ─────────────────────────────────────────────────────────────
all_rows:   list[dict] = []
page_stats: dict[str, dict] = {}

for sheet_name in page_sheets:
    ws     = wb[sheet_name]
    counts = {s: 0 for s in ALL_STATUSES}
    found  = 0

    for row_obj in ws.iter_rows(min_row=DATA_START_ROW, max_col=DATA_COL_COUNT):
        src_row = row_obj[0].row
        vals = [str(cell.value or "").strip() if cell.value is not None else ""
                for cell in row_obj]
        while len(vals) < DATA_COL_COUNT:
            vals.append("")

        if len(vals) < 8:
            continue

        element_id, page, typ, label, value, acc_id, status, new_status = vals[:8]
        ai_suggestion = vals[8] if len(vals) > 8 else ""

        if all(v == "" for v in vals[:8]):
            continue
        if status not in ALL_STATUSES:
            continue
        if not new_status:
            new_status = get_new_status_default(status)

        all_rows.append({
            "element_id":    element_id,
            "page":          page or sheet_name,
            "type":          typ,
            "label":         label,
            "value":         value,
            "acc_id":        acc_id,
            "status":        status,
            "new_status":    new_status,
            "ai_suggestion": ai_suggestion,
            "src_sheet":     sheet_name,
            "src_row":       src_row,
        })
        counts[status] += 1
        found += 1

    page_stats[sheet_name] = counts
    print(f"   ✓ {sheet_name:25s} → "
          f"Missing:{counts[sh.STATUS_MISSING]:3d}  "
          f"Undefined:{counts[sh.STATUS_UNDEFINED]:3d}  "
          f"Duplicate:{counts[sh.STATUS_DUPLICATE]:3d}  "
          f"Unique:{counts[sh.STATUS_UNIQUE]:3d}  "
          f"(toplam {found})")

# ─────────────────────────────────────────────────────────────
# 3. "Data" sheet'i oluştur
# ─────────────────────────────────────────────────────────────
if "Data" in wb.sheetnames: del wb["Data"]
wd = wb.create_sheet("Data", 0)

COLS_D   = ["Element ID", "Page", "Type", "Label / Text", "Value",
             "Resource ID", "Status", "AI Suggestion"]
WIDTHS_D = [22, 20, 16, 26, 18, 32, 14, 45]

wd.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS_D))
c = wd.cell(row=1, column=1, value="Accessibility ID — Ham Veri (Tüm Sayfalar)")
c.font = font(bold=True, color="FFFFFF", size=13)
c.fill = fill("375623"); c.alignment = CENTER; c.border = BORDER
wd.row_dimensions[1].height = 26

for ci, col_name in enumerate(COLS_D, 1):
    c = wd.cell(row=2, column=ci, value=col_name)
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(AI_COLOR["hdr"] if col_name == "AI Suggestion" else "2C2C2A")
    c.alignment = CENTER; c.border = BORDER
wd.row_dimensions[2].height = 18
wd.freeze_panes = "A3"

for idx, row in enumerate(all_rows):
    dr      = idx + 3
    pal     = PALETTE.get(row["status"], PALETTE[sh.STATUS_MISSING])
    r_fill  = fill(pal["row"] if idx % 2 == 0 else pal["alt"])
    ai_fill = fill(AI_COLOR["row"] if idx % 2 == 0 else AI_COLOR["alt"])
    values  = [row["element_id"], row["page"], row["type"], row["label"],
               row["value"], row["acc_id"], row["status"], row["ai_suggestion"]]

    for ci, val in enumerate(values, 1):
        c = wd.cell(row=dr, column=ci, value=val)
        c.border = BORDER
        if ci == 8:    # AI Suggestion
            c.fill = ai_fill
            c.font = Font(size=8, color=AI_COLOR["txt"])
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        elif ci == 7:  # Status
            c.fill = r_fill
            c.font = font(bold=True, color=pal["txt"])
            c.alignment = CENTER
        elif ci == 1:  # Element ID
            c.fill = r_fill
            c.font = font(bold=True)
            c.alignment = CENTER
        else:
            c.fill = r_fill
            c.font = font()
            c.alignment = LEFT
    wd.row_dimensions[dr].height = 55

for ci, w in enumerate(WIDTHS_D, 1):
    wd.column_dimensions[get_column_letter(ci)].width = w

print(f"\n✅ Data sheet oluşturuldu — {len(all_rows)} satır")

# ─────────────────────────────────────────────────────────────
# 4. "Summary" sheet'i oluştur
# ─────────────────────────────────────────────────────────────
if "Summary" in wb.sheetnames: del wb["Summary"]
ws_sum = wb.create_sheet("Summary", 1)

totals      = {s: sum(page_stats[p][s] for p in page_sheets) for s in ALL_STATUSES}
grand_total = sum(totals.values())

ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
c = ws_sum.cell(row=1, column=1, value="Accessibility ID — Özet Rapor")
c.font = font(bold=True, color="FFFFFF", size=14)
c.fill = fill("1F3864"); c.alignment = CENTER; c.border = BORDER
ws_sum.row_dimensions[1].height = 30

METRIC_COLS = {
    sh.STATUS_MISSING:   (1, "C00000", "FFDAD6", "501313"),
    sh.STATUS_UNDEFINED: (2, "C55A11", "FCE4D6", "412402"),
    sh.STATUS_DUPLICATE: (3, "7B3F00", "FAEEDA", "3B1F00"),
    sh.STATUS_UNIQUE:    (4, "375623", "E2EFDA", "173404"),
}

for status, (col, hdr_c, bg_c, txt_c) in METRIC_COLS.items():
    c = ws_sum.cell(row=3, column=col, value=status)
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(hdr_c); c.alignment = CENTER; c.border = BORDER
    ws_sum.row_dimensions[3].height = 20

    c = ws_sum.cell(row=4, column=col, value=totals[status])
    c.font = font(bold=True, color=txt_c, size=22)
    c.fill = fill(bg_c); c.alignment = CENTER; c.border = BORDER
    ws_sum.row_dimensions[4].height = 36

    pct = f"%{round(totals[status] / grand_total * 100) if grand_total else 0}"
    c = ws_sum.cell(row=5, column=col, value=pct)
    c.font = font(color=txt_c)
    c.fill = fill(bg_c); c.alignment = CENTER; c.border = BORDER
    ws_sum.row_dimensions[5].height = 18

for col in range(1, 5):
    ws_sum.column_dimensions[get_column_letter(col)].width = 22

TABLE_COLS   = ["Sayfa", sh.STATUS_MISSING, sh.STATUS_UNDEFINED,
                sh.STATUS_DUPLICATE, sh.STATUS_UNIQUE, "Toplam"]
TABLE_COLORS = ["2C2C2A", "C00000", "C55A11", "7B3F00", "375623", "1F3864"]
WIDTHS_S     = [24, 14, 14, 14, 14, 12]
tbl_start    = 8

ws_sum.merge_cells(start_row=tbl_start - 1, start_column=1,
                   end_row=tbl_start - 1,   end_column=len(TABLE_COLS))
c = ws_sum.cell(row=tbl_start - 1, column=1, value="Sayfa Bazlı Dağılım")
c.font = font(bold=True, color="FFFFFF", size=11)
c.fill = fill("1F3864"); c.alignment = CENTER; c.border = BORDER
ws_sum.row_dimensions[tbl_start - 1].height = 22

for ci, (col_name, col_color) in enumerate(zip(TABLE_COLS, TABLE_COLORS), 1):
    c = ws_sum.cell(row=tbl_start, column=ci, value=col_name)
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(col_color); c.alignment = CENTER; c.border = BORDER
ws_sum.row_dimensions[tbl_start].height = 18
ws_sum.freeze_panes = f"A{tbl_start + 1}"

for idx, page_name in enumerate(page_sheets):
    row_num   = tbl_start + 1 + idx
    counts    = page_stats[page_name]
    row_total = sum(counts.values())
    row_bg    = "F1EFE8" if idx % 2 == 0 else "FFFFFF"
    txt_cols  = ["2C2C2A", "C00000", "C55A11", "7B3F00", "375623", "1F3864"]
    values    = [page_name, counts[sh.STATUS_MISSING], counts[sh.STATUS_UNDEFINED],
                 counts[sh.STATUS_DUPLICATE], counts[sh.STATUS_UNIQUE], row_total]

    for ci, (val, txt_c) in enumerate(zip(values, txt_cols), 1):
        c = ws_sum.cell(row=row_num, column=ci, value=val)
        c.fill = fill(row_bg); c.border = BORDER
        if ci == 1:
            c.font = font(bold=True, color=txt_c); c.alignment = LEFT
        else:
            c.font = font(bold=(val > 0 and ci != 6), color=txt_c)
            c.alignment = CENTER
    ws_sum.row_dimensions[row_num].height = 16

total_row = tbl_start + 1 + len(page_sheets)
ws_sum.row_dimensions[total_row].height = 20
totals_values = ["TOPLAM", totals[sh.STATUS_MISSING], totals[sh.STATUS_UNDEFINED],
                 totals[sh.STATUS_DUPLICATE], totals[sh.STATUS_UNIQUE], grand_total]
for ci, (val, col_color) in enumerate(zip(totals_values, TABLE_COLORS), 1):
    c = ws_sum.cell(row=total_row, column=ci, value=val)
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill(col_color)
    c.alignment = LEFT if ci == 1 else CENTER
    c.border = BORDER

for ci, w in enumerate(WIDTHS_S, 1):
    ws_sum.column_dimensions[get_column_letter(ci)].width = w

print(f"✅ Summary sheet oluşturuldu — {len(page_sheets)} sayfa, {grand_total} element")

# ─────────────────────────────────────────────────────────────
# 5. "Task" sheet'i oluştur
# ─────────────────────────────────────────────────────────────
if "Task" in wb.sheetnames: del wb["Task"]
wt = wb.create_sheet("Task", 2)

task_rows = [r for r in all_rows if r["new_status"] == NS_WAITING]

COLS_T   = ["Element ID", "Page", "Type", "Label / Text", "Value",
             "Resource ID", "Status", "New Status", "AI Suggestion"]
COL_KEYS = ["element_id", "page", "type", "label", "value", "acc_id", "status"]
WIDTHS_T = [22, 20, 16, 26, 18, 32, 14, 28, 45]

wt.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS_T))
c = wt.cell(row=1, column=1,
            value=f"ID Eklenecek Elementler — Waiting Dev  ({len(task_rows)} adet)")
c.font = font(bold=True, color="FFFFFF", size=13)
c.fill = fill(NEW_STATUS_COLOR["hdr"]); c.alignment = CENTER; c.border = BORDER
wt.row_dimensions[1].height = 26

for ci, col_name in enumerate(COLS_T, 1):
    c = wt.cell(row=2, column=ci, value=col_name)
    c.font = font(bold=True, color="FFFFFF")
    if col_name == "AI Suggestion":  c.fill = fill(AI_COLOR["hdr"])
    elif col_name == "New Status":   c.fill = fill(NEW_STATUS_COLOR["hdr"])
    else:                            c.fill = fill("2C2C2A")
    c.alignment = CENTER; c.border = BORDER
wt.row_dimensions[2].height = 18
wt.freeze_panes = "A3"

for idx, row in enumerate(task_rows):
    tr      = idx + 3
    pal     = PALETTE.get(row["status"], PALETTE[sh.STATUS_MISSING])
    r_fill  = fill(pal["row"] if idx % 2 == 0 else pal["alt"])
    ns_fill = fill(NEW_STATUS_COLOR["row"] if idx % 2 == 0 else NEW_STATUS_COLOR["alt"])
    ai_fill = fill(AI_COLOR["row"] if idx % 2 == 0 else AI_COLOR["alt"])

    for ci, key in enumerate(COL_KEYS, 1):
        c = wt.cell(row=tr, column=ci, value=row[key])
        c.border = BORDER; c.fill = r_fill
        if ci == 7:    # Status
            c.font = font(bold=True, color=pal["txt"]); c.alignment = CENTER
        elif ci == 1:  # Element ID
            c.font = font(bold=True); c.alignment = CENTER
        else:
            c.font = font(); c.alignment = LEFT
    wt.row_dimensions[tr].height = 55

    # Sütun 8: New Status → kaynak hücreye formülle bağla
    c = wt.cell(row=tr, column=8,
                value=f"={sheet_ref(row['src_sheet'])}!{NEW_STATUS_COL_LTR}{row['src_row']}")
    c.border = BORDER; c.fill = ns_fill
    c.font = font(bold=True, color=NEW_STATUS_COLOR["txt"]); c.alignment = CENTER

    # Sütun 9: AI Suggestion → kaynak hücreye formülle bağla
    c = wt.cell(row=tr, column=9,
                value=f"={sheet_ref(row['src_sheet'])}!{AI_SUGGESTION_COL_LTR}{row['src_row']}")
    c.border = BORDER; c.fill = ai_fill
    c.font = Font(size=8, color=AI_COLOR["txt"])
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

for ci, w in enumerate(WIDTHS_T, 1):
    wt.column_dimensions[get_column_letter(ci)].width = w

print(f"✅ Task sheet oluşturuldu — {len(task_rows)} satır")

# ─────────────────────────────────────────────────────────────
# 6. Kaydet (safe_save: temp → atomic rename)
# ─────────────────────────────────────────────────────────────
sh.safe_save(wb, EXCEL_FILE)
print(f"\n📊 Dosya güncellendi: {EXCEL_FILE}")
print(f"   Sheet sırası: Data → Summary → Task → {' → '.join(page_sheets)}")
print(f"\n💡 Task sheet'indeki New Status ve AI Suggestion formülle bağlı.")
print(f"   Developer/QA kaynak sheet'te güncellediğinde Task otomatik yansır.")